"""Excel file adapter."""

from html import escape

from openpyxl import load_workbook

from app.adapters.file_adapter import FileAdapter


class ExcelAdapter(FileAdapter):
    def to_html_table(self, file_path: str) -> str:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        if not rows:
            return "<table></table>"

        header = rows[0]
        body_rows = rows[1:]

        html_parts = ["<table>", "<thead>", "<tr>"]
        html_parts.extend(f"<th>{self._format_cell(cell)}</th>" for cell in header)
        html_parts.extend(["</tr>", "</thead>", "<tbody>"])

        for row in body_rows:
            html_parts.append("<tr>")
            html_parts.extend(f"<td>{self._format_cell(cell)}</td>" for cell in row)
            html_parts.append("</tr>")

        html_parts.extend(["</tbody>", "</table>"])
        return "".join(html_parts)

    def _format_cell(self, cell: object) -> str:
        if cell is None:
            return ""
        return escape(str(cell))
